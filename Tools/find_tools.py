from openpyxl import load_workbook
import os
from langchain.tools import tool
from pydantic import BaseModel, Field
import time

def read_excel_sheet(sheet_name, file_path):
    workbook = load_workbook(filename=file_path)
    worksheet = workbook[sheet_name]

    def iterate_rows():
        for row in worksheet.iter_rows():
            row_str = '\t'.join(str(cell.value) for cell in row)
            yield row_str


    yield from iterate_rows()

def testA():
    if not hasattr(testA, 'output_iterator'):
        current_dir = os.path.dirname(__file__)
        file_path =  os.path.join(current_dir, '../Files/Test.xlsx')  # 替换为你的 Excel 文件路径
        sheet_name = '13-1'  # 替换为你要读取的 sheet 名称

        testA.output_iterator = read_excel_sheet(sheet_name, file_path)

    try:
        row = next(testA.output_iterator)
        return f'The content is: {row},Please take out the value according to the content and check if there are any parameters that have not been assigned a value. If there are parameters that are still not assigned a value, continue to search until all parameters are assigned a value.'
    except StopIteration:
        return "None"

@tool("Obtaining screen parameters",return_direct=False)
def find_params_yield_tool(test:str) -> str:
  '''Obtaining screen parameters from data requires repeated calls'''
  time.sleep(1)
  return testA()

    








