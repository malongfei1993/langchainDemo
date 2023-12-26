from openpyxl import load_workbook

def read_excel_sheet(sheet_name, file_path):
    workbook = load_workbook(filename=file_path)
    worksheet = workbook[sheet_name]

    total_words = 0
    result_str = ""

    for row in worksheet.iter_rows(max_row=10, values_only=True):
        row_str = '\t'.join(str(cell) for cell in row)
        result_str += row_str + '\n'
        total_words += len(row_str.split())

    result_str += "Total words: " + str(total_words)
    return result_str

# 调用示例
file_path = 'Files/Test.xlsx'  # 替换为你的 Excel 文件路径
sheet_name = '13-1'  # 替换为你要读取的 sheet 名称

# output = read_excel_sheet(sheet_name, file_path)
# print(output)